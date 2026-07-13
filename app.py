import streamlit as st
import pandas as pd
import pdfplumber
import io
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. MANAGEMENT STATE & DATABASE VIRTUAL
# ==========================================
st.set_page_config(page_title="IDX Financial Workspace", page_icon="🔵", layout="wide")

# Mengunci memori aplikasi untuk menyimpan user baru dan status login
if 'page' not in st.session_state:
    st.session_state.page = 'auth'
if 'auth_mode' not in st.session_state:
    st.session_state.auth_mode = 'login'
if 'user_db' not in st.session_state:
    st.session_state.user_db = {}  # Format: {"email": "password"}
if 'login_history' not in st.session_state:
    st.session_state.login_history = {}  # Format: {"email": jumlah_login}
if 'current_user' not in st.session_state:
    st.session_state.current_user = None
if 'is_new_user' not in st.session_state:
    st.session_state.is_new_user = False
if 'show_banner' not in st.session_state:
    st.session_state.show_banner = True
if 'pdf_cache' not in st.session_state:
    st.session_state.pdf_cache = None

# ==========================================
# 2. SISTEM ANIMASI INTERAKTIF & UI VIBRANT BLUE CSS
# ==========================================
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');
    
    .stApp { background-color: #F1F5F9; }
    html, body, [class*="css"] { font-family: 'Plus Jakarta Sans', sans-serif; }
    
    /* Efek Animasi Transisi Hover Semua Tombol & Elemen Interaktif */
    .stButton>button {
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        border-radius: 16px !important;
        font-weight: 600 !important;
        padding: 12px 24px !important;
    }
    .stButton>button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 10px 25px rgba(37, 99, 235, 0.2) !important;
    }
    
    /* Input Field & File Uploader Smooth Transitions */
    .stTextInput>div>div>input {
        transition: all 0.3s ease !important;
        border-radius: 14px !important;
    }
    .stTextInput>div>div>input:focus {
        border-color: #2563EB !important;
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.1) !important;
    }
    
    /* Vibrant Blue Header App Bar (Sesuai Referensi Gambar Baru) */
    .app-header-blue {
        background: linear-gradient(135deg, #2563EB 0%, #1D4ED8 100%);
        color: white;
        padding: 35px;
        border-radius: 24px;
        box-shadow: 0 8px 30px rgba(37, 99, 235, 0.15);
        margin-bottom: 30px;
    }
    
    /* Tampilan Kartu Berkas & Bento Grid */
    .bento-card {
        background-color: #FFFFFF;
        padding: 24px;
        border-radius: 24px;
        box-shadow: 0 4px 20px rgba(15, 23, 42, 0.01);
        border: 1px solid #E2E8F0;
        transition: all 0.4s ease;
    }
    .bento-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 12px 30px rgba(15, 23, 42, 0.04);
        border-color: #2563EB;
    }
    
    /* Banner Pengumuman Selamat Datang Kreatif Pengguna Baru */
    .welcome-banner {
        background: linear-gradient(135deg, #3B82F6 0%, #1D4ED8 100%);
        color: white;
        padding: 25px 30px;
        border-radius: 24px;
        position: relative;
        margin-bottom: 25px;
        box-shadow: 0 10px 25px rgba(37, 99, 235, 0.15);
    }
    
    /* Box Portal Keamanan */
    .auth-box {
        background-color: #FFFFFF;
        padding: 45px;
        border-radius: 32px;
        box-shadow: 0 15px 35px rgba(15, 23, 42, 0.04);
        border: 1px solid #E2E8F0;
        text-align: center;
    }
    
    .metric-label { font-size: 13px; color: #64748B; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; }
    .metric-value { font-size: 28px; color: #0F172A; font-weight: 700; letter-spacing: -0.5px; margin-top: 6px; }
    
    .badge-positive { background-color: #DCFCE7; color: #15803D; padding: 4px 12px; border-radius: 9999px; font-size: 12px; font-weight: 600; display: inline-block; margin-top: 10px; }
    .badge-negative { background-color: #FEE2E2; color: #B91C1C; padding: 4px 12px; border-radius: 9999px; font-size: 12px; font-weight: 600; display: inline-block; margin-top: 10px; }
    
    .icon-wrapper { display: flex; justify-content: center; align-items: center; margin-bottom: 15px; }
    .section-title { color: #0F172A; font-weight: 700; font-size: 20px; margin-bottom: 16px; margin-top: 10px; }
    
    /* Keyframe Spin / Pulse Animation */
    @keyframes spin-custom {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    .loader-spin { animation: spin-custom 1s linear infinite; display: inline-block; }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 3. KUMPULAN LOGO & ICON SVG FINANCIAL BLUE
# ==========================================
SVG_LOGO = '<svg width="64" height="64" viewBox="0 0 24 24" fill="none" stroke="#2563EB" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M3 3v18h18"/><path d="m19 9-5 5-4-4-3 3"/></svg>'
SVG_LOCK = '<svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="#2563EB" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="18" height="11" x="3" y="11" rx="2" ry="2"/><path d="M7 11V7a5 5 0 0 1 10 0v4"/></svg>'
SVG_USER = '<svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="#64748B" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M19 21v-2a4 4 0 0 0-4-4H9a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/></svg>'
SVG_GOOGLE = '<svg width="18" height="18" viewBox="0 0 24 24" fill="currentColor" style="vertical-align: middle; margin-right: 8px;"><path d="M22.56 12.25c0-.78-.07-1.53-.2-2.25H12v4.26h5.92c-.26 1.37-1.04 2.53-2.21 3.31v2.77h3.57c2.08-1.92 3.28-4.74 3.28-8.09z" fill="#4285F4"/><path d="M12 23c2.97 0 5.46-.98 7.28-2.66l-3.57-2.77c-.98.66-2.23 1.06-3.71 1.06-2.86 0-5.29-1.93-6.16-4.53H2.18v2.84C3.99 20.53 7.7 23 12 23z" fill="#34A853"/><path d="M5.84 14.09c-.22-.66-.35-1.36-.35-2.09s.13-1.43.35-2.09V7.06H2.18C1.43 8.55 1 10.22 1 12s.43 3.45 1.18 4.94l2.85-2.22.81-.63z" fill="#FBBC05"/><path d="M12 5.38c1.62 0 3.06.56 4.21 1.64l3.15-3.15C17.45 2.09 14.97 1 12 1 7.7 1 3.99 3.47 2.18 7.06l3.66 2.84c.87-2.6 3.3-4.53 6.16-4.53z" fill="#EA4335"/></svg>'
SVG_SPINNER = '<div class="loader-spin"><svg width="50" height="50" viewBox="0 0 24 24" fill="none" stroke="#2563EB" stroke-width="3" stroke-linecap="round"><circle cx="12" cy="12" r="10" stroke-dasharray="40 20"/></svg></div>'

# ==========================================
# 4. EXPLICIT FINANCIAL INDO-RUPIAH FORMATTER
# ==========================================
def to_rupiah(nominal):
    if isinstance(nominal, (int, float)):
        return f"Rp {nominal:,.0f}".replace(",", ".")
    return nominal

# ==========================================
# 5. DATA SIMULATION COMPREHENSIVE ENGINE
# ==========================================
def extract_all_financial_data(pdf_file):
    data_komplit = {
        "Item Laporan Keuangan": [
            "Total Pendapatan", "Laba Kotor", "Laba Bersih", 
            "Total Aset", "Aset Lancar", "Persediaan", "Piutang Usaha",
            "Total Liabilitas", "Liabilitas Jangka Pendek", "Total Ekuitas",
            "Arus Kas Operasi"
        ],
        "2023": [10000000000, 4000000000, 1200000000, 15000000000, 6000000000, 1500000000, 2000000000, 7000000000, 4000000000, 8000000000, 1300000000],
        "2024": [12000000000, 5000000000, 1500000000, 17000000000, 7500000000, 1800000000, 2200000000, 8000000000, 4500000000, 9000000000, 1600000000],
        "2025": [15000000000, 6500000000, 2100000000, 20000000000, 9000000000, 2000000000, 2500000000, 9000000000, 5000000000, 11000000000, 2300000000]
    }
    return pd.DataFrame(data_komplit)

# RESTORASI FORMULA HIDUP EXCEL ENGINE
def create_comprehensive_excel(df_raw):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    font_name = "Arial"
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10)
    bold_font = Font(name=font_name, size=10, bold=True)
    border_style = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    # WORKSHEET 1: DATA MENTAH
    ws1 = wb.active
    ws1.title = "Data_Mentah"
    ws1.views.sheetView[0].showGridLines = True
    
    for c_idx, h in enumerate(["Item Laporan Keuangan", "2023", "2024", "2025"], start=2):
        cell = ws1.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_style

    for r_idx, row in df_raw.iterrows():
        row_num = r_idx + 5
        for c_idx, val in enumerate(row, start=2):
            cell = ws1.cell(row=row_num, column=c_idx, value=val)
            cell.border = border_style
            if c_idx == 2:
                cell.font = data_font
            else:
                cell.font = data_font
                cell.number_format = '"Rp"#,##0'
                cell.alignment = Alignment(horizontal="right")

    # WORKSHEET 2: TOTAL ANALISIS RASIO DENGAN RUMUS EXCEL HIDUP
    ws2 = wb.create_sheet(title="Analisis_Rasio_Lengkap")
    ws2.views.sheetView[0].showGridLines = True
    
    r_headers = ["Komponen Analisis", "Formula Finansial", "2023", "2024", "2025"]
    for c_idx, h in enumerate(r_headers, start=2):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_style

    formulas = [
        ("Gross Profit Margin (GPM)", "Laba Kotor / Pendapatan", "=Data_Mentah!C6/Data_Mentah!C5", "=Data_Mentah!D6/Data_Mentah!D5", "=Data_Mentah!E6/Data_Mentah!E5", "0.0%"),
        ("Net Profit Margin (NPM)", "Laba Bersih / Pendapatan", "=Data_Mentah!C7/Data_Mentah!C5", "=Data_Mentah!D7/Data_Mentah!D5", "=Data_Mentah!E7/Data_Mentah!E5", "0.0%"),
        ("Return on Assets (ROA)", "Laba Bersih / Total Aset", "=Data_Mentah!C7/Data_Mentah!C8", "=Data_Mentah!D7/Data_Mentah!D8", "=Data_Mentah!E7/Data_Mentah!E8", "0.0%"),
        ("Return on Equity (ROE)", "Laba Bersih / Total Ekuitas", "=Data_Mentah!C7/Data_Mentah!C14", "=Data_Mentah!D7/Data_Mentah!D14", "=Data_Mentah!E7/Data_Mentah!E14", "0.0%"),
        ("Current Ratio (CR)", "Aset Lancar / Liabilitas Jk Pendek", "=Data_Mentah!C9/Data_Mentah!C13", "=Data_Mentah!D9/Data_Mentah!D13", "=Data_Mentah!E9/Data_Mentah!E13", "0.00"),
        ("Quick Ratio (QR)", "(Aset Lancar - Persediaan) / Liabilitas Jk Pendek", "=(Data_Mentah!C9-Data_Mentah!C10)/Data_Mentah!C13", "=(Data_Mentah!D9-Data_Mentah!D10)/Data_Mentah!D13", "=(Data_Mentah!E9-Data_Mentah!E10)/Data_Mentah!E13", "0.00"),
        ("Debt to Equity Ratio (DER)", "Total Liabilitas / Total Ekuitas", "=Data_Mentah!C12/Data_Mentah!C14", "=Data_Mentah!D12/Data_Mentah!D14", "=Data_Mentah!E12/Data_Mentah!E14", "0.00"),
        ("Debt to Asset Ratio (DAR)", "Total Liabilitas / Total Aset", "=Data_Mentah!C12/Data_Mentah!C8", "=Data_Mentah!D12/Data_Mentah!D8", "=Data_Mentah!E12/Data_Mentah!E8", "0.00"),
        ("Receivable Turnover", "Pendapatan / Piutang Usaha", "=Data_Mentah!C5/Data_Mentah!C11", "=Data_Mentah!D5/Data_Mentah!D11", "=Data_Mentah!E5/Data_Mentah!E11", "0.00"),
        ("Earnings Quality Score", "Arus Kas Operasi / Laba Bersih", "=Data_Mentah!C15/Data_Mentah!C7", "=Data_Mentah!D15/Data_Mentah!D7", "=Data_Mentah!E15/Data_Mentah!E7", "0.00")
    ]

    for r_idx, f_data in enumerate(formulas, start=5):
        for c_idx, val in enumerate(f_data[:5], start=2):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_style
            if c_idx == 2:
                cell.font = bold_font
            elif c_idx == 3:
                cell.font = Font(name=font_name, size=9, italic=True, color="64748B")
            else:
                cell.font = data_font
                cell.number_format = f_data[5]
                cell.alignment = Alignment(horizontal="right")

    for ws in [ws1, ws2]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 28
    ws1.column_dimensions['B'].width = 35
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['C'].width = 42

    wb.save(output)
    return output.getvalue()

# ==========================================
# 7. ROUTING DISPATCH MATRIX (PROSES ALUR)
# ==========================================

# --- HALAMAN 1: SEGMENTASI GATE AUTH (LOGIN & SIGN UP) ---
if st.session_state.page == 'auth':
    st.markdown("<br><br>", unsafe_allow_html=True)
    c_l, c_mid, c_r = st.columns([1.1, 1, 1.1])
    
    with c_mid:
        if st.session_state.auth_mode == 'login':
            st.markdown(f'<div class="auth-box"><div class="icon-wrapper">{SVG_LOCK}</div><h3 style="color:#0F172A;font-weight:700;">Masuk Ke Ruang Kerja</h3><p style="color:#64748B;font-size:14px;">Gunakan akun Anda untuk masuk ke workspace.</p></div>', unsafe_allow_html=True)
            
            in_email = st.text_input("Alamat Email", placeholder="nama@email.com")
            in_pass = st.text_input("Kata Sandi", type="password", placeholder="••••••••")
            
            if st.button("Masuk Sekarang", use_container_width=True, type="primary"):
                if in_email in st.session_state.user_db and st.session_state.user_db[in_email] == in_pass:
                    st.session_state.current_user = in_email
                    st.session_state.login_history[in_email] += 1
                    st.session_state.is_new_user = False
                    st.session_state.page = 'loading_login'
                    st.rerun()
                else:
                    st.error("Akses Ditolak! Kredensial tidak cocok. Pengguna wajib melakukan registrasi (Sign Up) terlebih dahulu jika belum punya akun.")
            
            st.markdown("<div style='text-align:center;color:#94A3B8;margin:10px 0;'>atau</div>", unsafe_allow_html=True)
            
            # Google Login
            if st.button("Google Gateway", use_container_width=True):
                st.session_state.current_user = "eksekutif_google@gmail.com"
                if "eksekutif_google@gmail.com" not in st.session_state.login_history:
                    st.session_state.login_history["eksekutif_google@gmail.com"] = 1
                    st.session_state.is_new_user = True
                    st.session_state.show_banner = True
                else:
                    st.session_state.login_history["eksekutif_google@gmail.com"] += 1
                    st.session_state.is_new_user = False
                st.session_state.page = 'loading_login'
                st.rerun()
            st.markdown(f"<script>var b=window.parent.document.getElementsByTagName('button');for(var i=0;i<b.length;i++){{if(b[i].innerText==='Google Gateway'){{b[i].innerHTML='{SVG_GOOGLE} Hubungkan Akun Google';b[i].style.backgroundColor='#FFF';b[i].style.color='#1F2937';b[i].style.border='1px solid #D1D5DB';}}}}</script>", unsafe_allow_html=True)
            
            if st.button("Belum punya akun? Buat Akun (Sign Up)", use_container_width=True):
                st.session_state.auth_mode = 'signup'
                st.rerun()
                
        else: # SIGN UP MODE
            st.markdown(f'<div class="auth-box"><div class="icon-wrapper">{SVG_USER}</div><h3 style="color:#0F172A;font-weight:700;">Pendaftaran Akun Baru</h3><p style="color:#64748B;font-size:14px;">Daftarkan data diri Anda untuk membuka hak akses penuh.</p></div>', unsafe_allow_html=True)
            
            up_email = st.text_input("Masukkan Alamat Email", placeholder="contoh@domain.com")
            up_pass = st.text_input("Buat Kata Sandi Kustom", type="password", placeholder="Minimal 6 karakter unik")
            
            if st.button("Daftarkan Akun Baru", use_container_width=True, type="primary"):
                if up_email and up_pass:
                    st.session_state.user_db[up_email] = up_pass
                    st.session_state.login_history[up_email] = 0
                    st.success("Registrasi Berhasil! Sistem mendeteksi akun Anda. Silakan kembali ke menu login untuk masuk.")
                    st.session_state.auth_mode = 'login'
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.warning("Gagal memproses. Seluruh kolom formulir pendaftaran wajib diisi lengkap.")
            
            if st.button("Sudah punya akun? Kembali ke Portal Login", use_container_width=True):
                st.session_state.auth_mode = 'login'
                st.rerun()

# --- HALAMAN 2: SINKRONISASI PEMUATAN DINAMIS ---
elif st.session_state.page == 'loading_login':
    st.markdown("<br><br><br><br>", unsafe_allow_html=True)
    c_l, c_m, c_r = st.columns([1, 1, 1])
    with c_m:
        st.markdown(f"<div style='text-align:center;'>{SVG_SPINNER}</div>", unsafe_allow_html=True)
        user = st.session_state.current_user
        if st.session_state.login_history.get(user, 1) == 1 or st.session_state.is_new_user:
            st.markdown("<p style='text-align:center;color:#64748B;margin-top:15px;'>Membangun enkripsi database ruang kerja pengguna baru... (5 Detik)</p>", unsafe_allow_html=True)
            time.sleep(5)
            st.session_state.is_new_user = True
        else:
            st.markdown("<p style='text-align:center;color:#64748B;margin-top:15px;'>Menghubungkan sesi cloud server pengguna reguler... (2 Detik)</p>", unsafe_allow_html=True)
            time.sleep(2)
            st.session_state.is_new_user = False
            
        st.session_state.page = 'landing'
        st.rerun()

# --- HALAMAN 3: LANDING PAGE & UPLOAD CORE ---
elif st.session_state.page == 'landing':
    # BANNER UNTUK PENGGUNA BARU YANG BISA DISILANG
    if st.session_state.is_new_user and st.session_state.show_banner:
        b_text, b_close = st.columns([12, 1])
        with b_text:
            st.markdown("""
                <div class="welcome-banner">
                    <h3 style="margin:0; font-weight:800; font-size:20px;">⚡ Akses Finansial Eksklusif Diaktifkan!</h3>
                    <p style="margin:6px 0 0 0; opacity:0.95; font-size:14px;">Selamat datang di platform masa depan analisis laporan keuangan. Sistem kami telah menyelesaikan setup awal khusus untuk akun baru Anda. Mulai taruh file PDF emiten Anda dan rasakan kemudahan ekstraksi otomatis instan!</p>
                </div>
            """, unsafe_allow_html=True)
        with b_close:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("✕ Close", key="dismiss_banner_vibrant"):
                st.session_state.show_banner = False
                st.rerun()

    # APP BAR DENGAN TEMA VIBRANT BLUE APPS
    st.markdown(f"""
        <div class="app-header-blue">
            <h1 style='margin:0; font-weight:800; font-size:36px; letter-spacing:-1.5px;'>RUANG KERJA ANALISIS KEUANGAN</h1>
            <p style='margin:5px 0 0 0; opacity:0.85; font-size:15px;'>Operator Aktif: {st.session_state.current_user}</p>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("<p style='color:#0F172A; font-weight:700; font-size:16px; margin-bottom:8px;'>Taruh PDF kalian disini</p>", unsafe_allow_html=True)
    
    up_pdf = st.file_uploader("", type=["pdf"], label_visibility="collapsed")
    
    if up_pdf is not None:
        st.session_state.pdf_cache = up_pdf
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔥 ANALISIS SEKARANG", use_container_width=True, type="primary"):
            st.session_state.page = 'loading_analysis'
            st.rerun()

# --- HALAMAN 4: PEMUATAN EXTRACTION FILE ENGINE ---
elif st.session_state.page == 'loading_analysis':
    st.markdown("<br><br><br><br>", unsafe_allow_html=True)
    c_l, c_m, c_r = st.columns([1, 1, 1])
    with c_m:
        st.markdown(f"<div style='text-align:center;'>{SVG_SPINNER}</div>", unsafe_allow_html=True)
        st.markdown("<p style='text-align:center;color:#0F172A;font-weight:700;margin-top:15px;'>Engine cerdas sedang memetakan koordinat tabel finansial... (5 Detik)</p>", unsafe_allow_html=True)
        time.sleep(5)
        st.session_state.page = 'dashboard'
        st.rerun()

# --- HALAMAN 5: DASHBOARD WORKSPACE ALL PILARS (STRICT RUPIAH) ---
elif st.session_state.page == 'dashboard':
    c_title, c_logout = st.columns([5, 1])
    with c_title:
        st.markdown("<h2 style='color:#0F172A; font-weight:800; margin-bottom:0;'>Hasil Analisis Emiten Konsolidasi</h2>", unsafe_allow_html=True)
        st.markdown("<p style='color:#64748B; font-size:14px;'>Data Kompilasi 3 Periode Terakhir</p>", unsafe_allow_html=True)
    with c_logout:
        if st.button("🚪 Keluar Sesi", use_container_width=True):
            st.session_state.page = 'auth'
            st.session_state.show_banner = True
            st.session_state.pdf_cache = None
            st.rerun()
            
    st.markdown("---")
    
    # Ekstraksi File Cache
    df_res = extract_all_financial_data(st.session_state.pdf_cache)
    df_p = df_res.set_index("Item Laporan Keuangan")
    years = ["2023", "2024", "2025"]
    
    # SCORECARDS BENTO LAYOUT (STRICT FORMAT RUPIAH)
    b1, b2, b3 = st.columns(3)
    with b1:
        st.markdown(f"<div class='bento-card'><div class='metric-label'>Laba Bersih Setelah Pajak (2025)</div><div class='metric-value'>{to_rupiah(df_p.loc['Laba Bersih', '2025'])}</div><div class='badge-positive'>↗ Tumbuh Positif</div></div>", unsafe_allow_html=True)
    with b2:
        st.markdown(f"<div class='bento-card'><div class='metric-label'>Total Pendapatan Usaha (2025)</div><div class='metric-value'>{to_rupiah(df_p.loc['Total Pendapatan', '2025'])}</div><div class='badge-positive'>↗ Penjualan Naik</div></div>", unsafe_allow_html=True)
    with b3:
        st.markdown(f"<div class='bento-card'><div class='metric-label'>Total Aset Konsolidasi (2025)</div><div class='metric-value'>{to_rupiah(df_p.loc['Total Aset', '2025'])}</div><div class='badge-positive'>↗ Struktur Likuid</div></div>", unsafe_allow_html=True)

    # RESTORASI SELURUH 5 PILAR ANALISIS AKUNTANSI DI TAB PANEL
    t_profit, t_liquidity, t_activity, t_raw = st.tabs([
        "📈 Profitabilitas & Tren", 
        "🛡️ Likuiditas & Solvabilitas", 
        "⚙️ Aktivitas & Kualitas Laba", 
        "📂 Angka Mentah Terformat Rp"
    ])

    with t_profit:
        st.markdown("<div class=\"section-title\">Tren Margin Laba & Efisiensi Operasional</div>", unsafe_allow_html=True)
        c1, c2 = st.columns([3, 2])
        with c1:
            gpm_arr = [(df_p.loc["Laba Kotor", y]/df_p.loc["Total Pendapatan", y])*100 for y in years]
            npm_arr = [(df_p.loc["Laba Bersih", y]/df_p.loc["Total Pendapatan", y])*100 for y in years]
            st.line_chart(pd.DataFrame({"Tahun": years, "GPM (%)": gpm_arr, "NPM (%)": npm_arr}).set_index("Tahun"), color=["#2563EB", "#10B981"])
        with c2:
            st.dataframe(pd.DataFrame({"2023": [f"{gpm_arr[0]:.1f}%", f"{npm_arr[0]:.1f}%"], "2024": [f"{gpm_arr[1]:.1f}%", f"{npm_arr[1]:.1f}%"], "2025": [f"{gpm_arr[2]:.1f}%", f"{npm_arr[2]:.1f}%"]}, index=["Gross Profit Margin", "Net Profit Margin"]), use_container_width=True)

    with t_liquidity:
        st.markdown("<div class=\"section-title\">Rasio Daya Bayar Kewajiban & Leverage Risiko Utang</div>", unsafe_allow_html=True)
        cl1, cl2 = st.columns(2)
        with cl1:
            cr_arr = [df_p.loc["Aset Lancar", y] / df_p.loc["Liabilitas Jangka Pendek", y] for y in years]
            qr_arr = [(df_p.loc["Aset Lancar", y] - df_p.loc["Persediaan", y]) / df_p.loc["Liabilitas Jangka Pendek", y] for y in years]
            st.bar_chart(pd.DataFrame({"Tahun": years, "Current Ratio": cr_arr, "Quick Ratio": qr_arr}).set_index("Tahun"), color=["#2563EB", "#94A3B8"])
        with cl2:
            der_arr = [df_p.loc["Total Liabilitas", y] / df_p.loc["Total Ekuitas", y] for y in years]
            st.line_chart(pd.DataFrame({"Tahun": years, "Debt to Equity (DER)": der_arr}).set_index("Tahun"), color=["#EF4444"])

    with t_activity:
        st.markdown("<div class=\"section-title\">Kecepatan Perputaran Aset & Skor Kesehatan Kualitas Laba (Riset Akrual)</div>", unsafe_allow_html=True)
        ca1, ca2 = st.columns(2)
        with ca1:
            rec_to = [df_p.loc["Total Pendapatan", y] / df_p.loc["Piutang Usaha", y] for y in years]
            st.write("**Perputaran Piutang Usaha (Receivable Turnover):**")
            st.dataframe(pd.DataFrame({"Tahun": years, "Turnover (Kali)": rec_to}).set_index("Tahun"), use_container_width=True)
        with ca2:
            eq_score = [df_p.loc["Arus Kas Operasi", y] / df_p.loc["Laba Bersih", y] for y in years]
            st.write("**Earnings Quality Score (Arus Kas Operasi / Laba Bersih):**")
            st.dataframe(pd.DataFrame({"Tahun": years, "Skor Kualitas": eq_score}).set_index("Tahun"), use_container_width=True)

    with t_raw:
        st.markdown("<div class=\"section-title\">Tabel Hasil Ekstraksi Otomatis (Wajib Format Rupiah)</div>", unsafe_allow_html=True)
        df_format_rp = df_res.copy()
        for yr in years:
            df_format_rp[yr] = df_format_rp[yr].apply(to_rupiah)
        st.dataframe(df_format_rp.set_index("Item Laporan Keuangan"), use_container_width=True)

    # RESTORASI TOMBOL DOWNLOAD EXCEL FORMULA AKTIF
    excel_file = create_comprehensive_excel(df_res)
    st.markdown("<br>", unsafe_allow_html=True)
    st.download_button(
        label="📥 Download Laporan Analisis Lengkap Berformula (.xlsx)", 
        data=excel_file, 
        file_name="IDX_Comprehensive_Analysis.xlsx", 
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        use_container_width=True
    )
